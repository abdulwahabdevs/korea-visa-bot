"""
Excel writer — produces the final results spreadsheet.
"""

from __future__ import annotations
import logging
from datetime import datetime
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Status colour map  (background fill hex, no #)
# Each status gets a DISTINCT colour so admins can scan the Excel at a glance.
STATUS_FILLS: dict[str, str] = {
    # Approved family — greens
    "APPROVED":        "C6EFCE",   # light green
    "ISSUED":          "C6EFCE",   # light green
    "USED":            "A9D18E",   # darker green (visa used, entered Korea)
    # Pending family — yellows / blues
    "PENDING":         "FFEB9C",   # yellow
    "RECEIVED":        "FFF2CC",   # lighter yellow (just received, no review)
    "UNDER_REVIEW":    "DDEBF7",   # light blue (actively being reviewed)
    "SUPPLEMENT_DONE": "D6E4F0",   # blue-grey (docs submitted, back in queue)
    # Supplement requested — orange (action required!)
    "SUPPLEMENT":      "F4B084",   # orange — stands out as "action needed"
    # Rejected / returned — reds
    "REJECTED":        "FFC7CE",   # red
    "RETURNED":        "F8CBAD",   # salmon/coral (returned by consulate)
    # Withdrawn / cancelled — greys
    "WITHDRAWN":       "D9D9D9",   # grey
    "CANCELLED":       "BFBFBF",   # darker grey (신청취소)
    # Other
    "NOT_FOUND":       "EDEDED",   # light grey
    "ERROR":           "E2BFDB",   # mauve/purple (distinct from red statuses)
    "UNKNOWN":         "FFFFFF",   # white
}

HEADERS = [
    "#",
    "신청번호 (Receipt No)",
    "피초청자 (Name)",
    "여권번호 (Passport)",
    "생년월일 (DOB)",
    "진행상태 (Status EN)",
    "진행상태(한국어) (Status KO)",
    "체류자격 (Visa Type)",
    "신청일자 (App Date)",
    "보완/불허 사유 (Reason)",
    "확인일자 (Checked)",
]

HISTORY_HEADERS = [
    "여권번호 (Passport)",
    "피초청자 (Name)",
    "생년월일 (DOB)",
    "신청번호 (Receipt)",
    "유형 (Type)",
    "진행상태 (Status EN)",
    "진행상태(한국어) (Status KO)",
    "체류자격 (Visa Type)",
    "신청일자 (App Date)",
    "사유 (Reason)",
    "최초 확인 (First Seen)",
    "마지막 확인 (Last Checked)",
    "상태 변경일 (Status Changed At)",
]


def _thin_border() -> Border:
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def _write_header_row(ws, headers: list[str], col_widths: list[int]) -> None:
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1F3864")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align
        cell.border = _thin_border()
    ws.row_dimensions[1].height = 32
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"


def write_results(rows: List[dict], output_path: str | Path) -> Path:
    """
    Write a list of result dicts to an Excel file.

    Each dict should have keys:
        index, receipt, name, passport, dob,
        status_en, status_ko, visa_type, app_date, reason, checked_at

    Returns the path of the written file.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Visa Status"

    col_widths = [5, 16, 36, 14, 12, 14, 22, 22, 14, 30, 20]
    _write_header_row(ws, HEADERS, col_widths)

    for row_idx, rec in enumerate(rows, start=2):
        if rec is None:
            continue
        status_en = rec.get("status_en", "UNKNOWN")
        fill_hex  = STATUS_FILLS.get(status_en, "FFFFFF")
        row_fill  = PatternFill("solid", fgColor=fill_hex)
        row_align = Alignment(vertical="center")

        values = [
            rec.get("index", row_idx - 1),
            rec.get("receipt", ""),
            rec.get("name", ""),
            rec.get("passport", ""),
            rec.get("dob", ""),
            status_en,
            rec.get("status_ko", ""),
            rec.get("visa_type", ""),
            rec.get("app_date", ""),
            rec.get("reason", ""),
            rec.get("checked_at", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = row_fill
            cell.alignment = row_align
            cell.border    = _thin_border()
            if col_idx == 6:  # Status EN — bold
                cell.font = Font(bold=True)

    wb.save(output_path)
    logger.info("Results written to %s (%d rows)", output_path, len(rows))
    return output_path


def write_history_export(rows: List[dict], output_path: str | Path) -> Path:
    """
    Write a full status-history export with change-tracking columns.

    Each dict comes from db.database.get_all_status_history() and contains:
        passport, full_name, dob, receipt, check_type,
        status_en, status_ko, visa_type, app_date, reason,
        first_seen_at, last_checked, status_changed_at
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Status History"

    col_widths = [16, 30, 12, 16, 12, 14, 26, 22, 14, 30, 20, 20, 22]
    _write_header_row(ws, HISTORY_HEADERS, col_widths)

    for row_idx, rec in enumerate(rows, start=2):
        status_en = rec.get("status_en", "UNKNOWN")
        fill_hex  = STATUS_FILLS.get(status_en, "FFFFFF")
        row_fill  = PatternFill("solid", fgColor=fill_hex)
        row_align = Alignment(vertical="center")

        # Highlight rows where status changed recently
        changed_at = rec.get("status_changed_at") or ""

        values = [
            rec.get("passport", ""),
            rec.get("full_name", ""),
            rec.get("dob", ""),
            rec.get("receipt", ""),
            rec.get("check_type", ""),
            status_en,
            rec.get("status_ko", ""),
            rec.get("visa_type", ""),
            rec.get("app_date", ""),
            rec.get("reason", ""),
            rec.get("first_seen_at", ""),
            rec.get("last_checked", ""),
            changed_at,
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = row_fill
            cell.alignment = row_align
            cell.border    = _thin_border()
            if col_idx == 6:
                cell.font = Font(bold=True)
            # Bold the status-changed-at column if it has a value
            if col_idx == 13 and changed_at:
                cell.font = Font(bold=True, color="C00000")

    wb.save(output_path)
    logger.info("History export written to %s (%d rows)", output_path, len(rows))
    return output_path
