"""
Excel reader with smart header detection and per-field validation.

Supports any column order, Korean or English headers,
extra columns, and various DOB date formats.

Validation rules applied to every row before any scraping:
  passport  — 1-2 uppercase letters + 7-8 digits  (e.g. FA4166021)
  receipt   — exactly 10 digits                    (e.g. 5677400001)
  name      — 2+ chars, letters/spaces/hyphens only
  dob       — YYYYMMDD after normalisation; year 1920-2015, month 1-12, day 1-31
  duplicates — same passport (+ receipt for e-visa) are flagged and skipped
"""

from __future__ import annotations
import re
import logging
from dataclasses import dataclass, field as dc_field
from typing import Optional
from pathlib import Path
import openpyxl

logger = logging.getLogger(__name__)


# ── Data structures ───────────────────────────────────────────────────────────

@dataclass
class StudentRecord:
    row_num: int
    receipt: str
    name:    str
    passport: str
    dob: str        # normalised to YYYYMMDD


@dataclass
class InvalidRow:
    """A row that failed validation — reported to admin before any scraping."""
    row_num:  int
    passport: str
    name:     str
    receipt:  str
    dob_raw:  str
    errors:   list = dc_field(default_factory=list)   # human-readable reasons


@dataclass
class ReadResult:
    """Return value of read_students_validated()."""
    valid:    list   # list[StudentRecord]
    invalid:  list   # list[InvalidRow]
    warnings: list   # list[str]  e.g. duplicate notices


# ── Header keyword mapping ────────────────────────────────────────────────────

HEADER_KEYWORDS: dict = {
    "receipt":  ["신청번호", "접수번호", "invitee_seq", "invitee seq", "seq no",
                 "receipt", "app no", "application no", "application", "ariza", "reg no"],
    "name":     ["피초청자", "성명", "영문성명", "name", "full name", "ism"],
    "passport": ["여권번호", "여권", "passport", "pass no", "pasport"],
    "dob":      ["생년월일", "생년", "birthdate", "birth", "dob", "date of birth", "birthday", "tug"],
}


def _match_header(cell_value: str) -> Optional[str]:
    """Return the field key if the cell matches a known header keyword."""
    val = str(cell_value).strip().lower()
    for field_key, keywords in HEADER_KEYWORDS.items():
        for kw in keywords:
            if kw in val:
                return field_key
    return None


# ── Field normalisers ─────────────────────────────────────────────────────────

def _normalise_dob(raw) -> str:
    """
    Convert various date formats to YYYYMMDD.

    Handles: 1998.12.17 / 1998-12-17 / 19981217 / 17/12/1998
             Excel serial integers (e.g. 36142)
    """
    if raw is None:
        return ""

    # Excel serial date integer
    if isinstance(raw, (int, float)):
        from datetime import date, timedelta
        epoch = date(1899, 12, 30)
        try:
            d = epoch + timedelta(days=int(raw))
            return d.strftime("%Y%m%d")
        except Exception:
            return str(raw)

    s = str(raw).strip()

    # Already YYYYMMDD
    if re.fullmatch(r"\d{8}", s):
        return s

    # YYYY.MM.DD / YYYY-MM-DD / YYYY/MM/DD
    m = re.match(r"(\d{4})[.\-/](\d{2})[.\-/](\d{2})", s)
    if m:
        return m.group(1) + m.group(2) + m.group(3)

    # DD/MM/YYYY or DD.MM.YYYY
    m = re.match(r"(\d{2})[.\-/](\d{2})[.\-/](\d{4})", s)
    if m:
        return m.group(3) + m.group(2) + m.group(1)

    # Fallback: strip non-digits
    digits = re.sub(r"\D", "", s)
    if len(digits) == 8:
        return digits

    return s


def _normalise_passport(raw: str) -> str:
    """Strip spaces, convert to uppercase."""
    return re.sub(r"\s", "", raw).strip().upper()


def _normalise_receipt(raw: str) -> str:
    """Strip all whitespace."""
    return re.sub(r"\s", "", raw).strip()


def _normalise_name(raw: str) -> str:
    """Collapse whitespace, uppercase."""
    return re.sub(r"\s+", " ", raw.strip()).upper()


# ── Per-field validators ──────────────────────────────────────────────────────

def _validate_passport(value: str) -> list:
    """Return list of error strings (empty = valid)."""
    if not value:
        return ["Passport bo'sh"]
    if not re.fullmatch(r"[A-Z]{1,2}\d{7,8}", value):
        return [f"Passport formati noto'g'ri: '{value}' (kutiladi: FA4166021)"]
    return []


def _validate_receipt(value: str) -> list:
    """Return list of error strings (empty = valid)."""
    if not value:
        return ["Ariza raqami bo'sh"]
    if not re.fullmatch(r"\d{10}", value):
        return [f"Ariza raqami formati noto'g'ri: '{value}' (kutiladi: 10 ta raqam)"]
    return []


def _validate_name(value: str) -> list:
    """Return list of error strings (empty = valid)."""
    if not value or len(value) < 2:
        return ["Ism bo'sh yoki juda qisqa"]
    if not re.fullmatch(r"[A-Z\s\-']+", value):
        return [f"Ismda noto'g'ri belgilar: '{value}' (faqat lotin harflari)"]
    return []


def _validate_dob(value: str) -> list:
    """Return list of error strings (empty = valid). Expects YYYYMMDD."""
    if not value or not re.fullmatch(r"\d{8}", value):
        return [f"Tug'ilgan sana formati noto'g'ri: '{value}' (kutiladi: YYYYMMDD)"]
    year  = int(value[:4])
    month = int(value[4:6])
    day   = int(value[6:])
    errors = []
    if not (1920 <= year <= 2015):
        errors.append(f"Tug'ilgan yil {year} mumkin emas (1920–2015)")
    if not (1 <= month <= 12):
        errors.append(f"Oy {month} noto'g'ri (1–12)")
    if not (1 <= day <= 31):
        errors.append(f"Kun {day} noto'g'ri (1–31)")
    return errors


# ── Core reader ───────────────────────────────────────────────────────────────

def _detect_headers(ws) -> tuple:
    """
    Scan first 10 rows to find the header row.
    Handles files where data starts from any column (A, B, C…).
    Returns (col_map, header_row_idx) where col_map values are
    0-based indices into each row tuple.
    """
    col_map: dict = {}
    header_row_idx = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
        candidate: dict = {}
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            fk = _match_header(cell)
            if fk and fk not in candidate:
                candidate[fk] = col_idx
        # Accept this row as the header if it contains at least 2 known fields
        if len(candidate) >= 2:
            col_map = candidate
            header_row_idx = row_idx
            break

    return col_map, header_row_idx


def read_students(path, require_receipt: bool = True) -> list:
    """
    Legacy interface — returns list[StudentRecord].
    Validates fields and SKIPS invalid rows silently (logs warnings).
    Use read_students_validated() to get the full report.
    """
    result = read_students_validated(path, require_receipt=require_receipt)
    if result.invalid:
        logger.warning(
            "%d invalid rows skipped from %s: %s",
            len(result.invalid),
            Path(path).name,
            [(r.row_num, r.errors) for r in result.invalid],
        )
    return result.valid


def read_students_validated(path, require_receipt: bool = True) -> "ReadResult":
    """
    Read an Excel file, validate every field, and return a ReadResult with:
      .valid    — StudentRecord list ready for the scraper
      .invalid  — InvalidRow list with per-row error messages (shown to admin)
      .warnings — duplicate / structural notices

    Args:
        path:            Path to the .xlsx / .xls file.
        require_receipt: True = E-Visa mode (receipt column mandatory, used to
                         anchor rows and validated as 10-digit number).
                         False = Diplomatic mode (receipt column optional).

    Raises:
        ValueError:        if required columns are missing entirely.
        FileNotFoundError: if the path does not exist.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    col_map, header_row_idx = _detect_headers(ws)

    if require_receipt and "receipt" not in col_map:
        raise ValueError(
            "Ariza raqami (receipt) ustuni topilmadi.\n"
            "Ustun sarlavhasi quyidagilardan biri bo'lishi kerak: "
            "신청번호, 접수번호, receipt, app no\n\n"
            "Agar Diplomatic Office tekshiruvi bo'lsa, 'Diplomatic Office' "
            "turini tanlang — u ariza raqamisiz ishlaydi."
        )
    if "passport" not in col_map:
        raise ValueError(
            "Pasport raqami ustuni topilmadi.\n"
            "Ustun sarlavhasi quyidagilardan biri bo'lishi kerak: "
            "여권번호, passport, pass no"
        )
    if "name" not in col_map:
        raise ValueError(
            "Ism ustuni topilmadi.\n"
            "Ustun sarlavhasi quyidagilardan biri bo'lishi kerak: "
            "피초청자, 성명, name, full name"
        )
    if "dob" not in col_map:
        raise ValueError(
            "Tug'ilgan sana ustuni topilmadi.\n"
            "Ustun sarlavhasi quyidagilardan biri bo'lishi kerak: "
            "생년월일, dob, date of birth, birthday"
        )

    logger.info("Header at row %d: %s", (header_row_idx or 0) + 1, col_map)

    valid:    list = []
    invalid:  list = []
    warnings: list = []
    data_start = (header_row_idx or 0) + 2  # 1-indexed

    seen_passports: dict = {}   # passport → first row_num (duplicate detection)
    seen_receipts:  dict = {}   # receipt  → first row_num (for e-visa duplicates)

    for row_num, row in enumerate(
        ws.iter_rows(min_row=data_start, values_only=True), start=data_start
    ):
        def _get(field_key: str) -> str:
            idx = col_map.get(field_key)
            if idx is None or idx >= len(row):
                return ""
            val = row[idx]
            return str(val).strip() if val is not None else ""

        raw_receipt  = _get("receipt")
        raw_passport = _get("passport")
        raw_name     = _get("name")
        raw_dob      = row[col_map["dob"]] if col_map.get("dob", 99) < len(row) else None

        # ── Skip fully blank rows ─────────────────────────────────────────
        anchor = raw_receipt if require_receipt else raw_passport
        if not anchor:
            continue

        # ── Normalise ─────────────────────────────────────────────────────
        passport = _normalise_passport(raw_passport)
        receipt  = _normalise_receipt(raw_receipt)
        name     = _normalise_name(raw_name)
        dob      = _normalise_dob(raw_dob)
        dob_raw_str = str(raw_dob).strip() if raw_dob is not None else ""

        # ── Validate ──────────────────────────────────────────────────────
        errors: list = []
        errors += _validate_passport(passport)
        errors += _validate_name(name)
        errors += _validate_dob(dob)
        if require_receipt:
            errors += _validate_receipt(receipt)

        # ── Duplicate detection ───────────────────────────────────────────
        if passport and passport in seen_passports:
            dup_row = seen_passports[passport]
            warnings.append(
                f"Takrорiy passport {passport}: satr {dup_row} va {row_num} — "
                f"satr {row_num} o'tkazib yuborildi"
            )
            # Mark as invalid with a clear reason (not a data error, but still skipped)
            invalid.append(InvalidRow(
                row_num=row_num, passport=passport, name=name,
                receipt=receipt, dob_raw=dob_raw_str,
                errors=[f"Takroriy passport (avval satr {dup_row}da ko'rilgan)"],
            ))
            continue
        if require_receipt and receipt and receipt in seen_receipts:
            dup_row = seen_receipts[receipt]
            warnings.append(
                f"Takroriy ariza raqami {receipt}: satr {dup_row} va {row_num} — "
                f"satr {row_num} o'tkazib yuborildi"
            )
            invalid.append(InvalidRow(
                row_num=row_num, passport=passport, name=name,
                receipt=receipt, dob_raw=dob_raw_str,
                errors=[f"Takroriy ariza raqami (avval satr {dup_row}da ko'rilgan)"],
            ))
            continue

        if errors:
            invalid.append(InvalidRow(
                row_num=row_num, passport=raw_passport,
                name=raw_name, receipt=raw_receipt,
                dob_raw=dob_raw_str, errors=errors,
            ))
            logger.warning("Row %d invalid: %s", row_num, errors)
            continue

        # ── Valid row ─────────────────────────────────────────────────────
        seen_passports[passport] = row_num
        if require_receipt:
            seen_receipts[receipt] = row_num

        valid.append(StudentRecord(
            row_num=row_num,
            receipt=receipt,
            name=name,
            passport=passport,
            dob=dob,
        ))

    logger.info(
        "Loaded from %s: %d valid, %d invalid, %d warnings",
        path.name, len(valid), len(invalid), len(warnings),
    )
    return ReadResult(valid=valid, invalid=invalid, warnings=warnings)
